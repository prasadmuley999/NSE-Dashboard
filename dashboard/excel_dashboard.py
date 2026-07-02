"""
Excel Dashboard Generator
Produces NSE_Dashboard.xlsx with formatting matching the reference file.
"""

import logging
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ─── Colour palette ──────────────────────────────────────────────────────────
C = {
    "bg_dark":    "1A1A2E",
    "bg_mid":     "16213E",
    "bg_card":    "0F3460",
    "green_str":  "00C851",
    "green_wk":   "43A047",
    "neutral":    "9E9E9E",
    "red_wk":     "E53935",
    "red_str":    "B71C1C",
    "gold":       "FFD700",
    "white":      "FFFFFF",
    "light_grey": "EEEEEE",
    "header_bg":  "0D47A1",
    "row_alt":    "E3F2FD",
}

SIGNAL_COLORS = {
    "🟢 Strong Bullish":  ("00C851", "FFFFFF"),
    "🟩 Weak Bullish":    ("43A047", "FFFFFF"),
    "⬜ Neutral":          ("9E9E9E", "FFFFFF"),
    "🟥 Weak Bearish":    ("E53935", "FFFFFF"),
    "🔴 Strong Bearish":  ("B71C1C", "FFFFFF"),
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color="000000", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")


def _border(style="thin") -> Border:
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _write(ws, row, col, value, bold=False, size=11, fg="000000",
           bg=None, h="center", v="center", wrap=False, border=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold, size=size, color=fg)
    cell.alignment = _align(h=h, v=v, wrap=wrap)
    if bg:
        cell.fill = _fill(bg)
    if border:
        cell.border = _border()
    return cell


def _header_row(ws, row, cols: list[str], bg=None):
    bg = bg or C["header_bg"]
    for i, label in enumerate(cols, 1):
        _write(ws, row, i, label, bold=True, fg=C["white"], bg=bg,
               border=True, wrap=True)


def _signal_cell(ws, row, col, label: str):
    bg, fg = SIGNAL_COLORS.get(label, ("9E9E9E", "FFFFFF"))
    _write(ws, row, col, label, bold=True, fg=fg, bg=bg, border=True)


def _set_col_widths(ws, widths: list[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _set_row_height(ws, row, h):
    ws.row_dimensions[row].height = h


# ─── Sheet 1: Dashboard ──────────────────────────────────────────────────────

def _sheet_dashboard(wb: Workbook, signals: pd.DataFrame, fii: pd.Series,
                     sector: pd.DataFrame, market_date: date, run_time: str):
    ws = wb.create_sheet("📊 Dashboard")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:P1")
    title = f"  NSE NIFTY 50  ·  POST-MARKET INTELLIGENCE DASHBOARD  ·  {market_date.strftime('%d %b %Y')}  {run_time}"
    _write(ws, 1, 1, title, bold=True, size=14, fg=C["gold"], bg=C["bg_dark"], h="left")
    _set_row_height(ws, 1, 30)

    # Signal summary bar
    counts = signals["signal_label"].value_counts()
    def _c(lbl): return counts.get(lbl, 0)

    summary_cols = [
        ("🟢 Strong\nBullish", str(_c("🟢 Strong Bullish")), C["green_str"]),
        ("🟩 Weak\nBullish",   str(_c("🟩 Weak Bullish")),   C["green_wk"]),
        ("⬜ Neutral",          str(_c("⬜ Neutral")),          C["neutral"]),
        ("🟥 Weak\nBearish",   str(_c("🟥 Weak Bearish")),   C["red_wk"]),
        ("🔴 Strong\nBearish", str(_c("🔴 Strong Bearish")), C["red_str"]),
    ]

    avg_del = signals["deliv_pct"].mean()
    fii_net = float(fii.get("fii_net", 0)) if not fii.empty else 0
    dii_net = float(fii.get("dii_net", 0)) if not fii.empty else 0

    extra_cols = [
        (f"Avg Del%\nToday\n{avg_del:.1f}%",       C["bg_card"]),
        (f"FII Net\n(₹ Cr)\n{fii_net:+.2f}",       C["bg_card"]),
        (f"DII Net\n(₹ Cr)\n{dii_net:+.2f}",       C["bg_card"]),
    ]

    col = 1
    for label, count, bg in summary_cols:
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        cell = ws.cell(row=3, column=col, value=f"{label}\n{count}")
        cell.font      = _font(bold=True, size=13, color=C["white"])
        cell.fill      = _fill(bg)
        cell.alignment = _align(wrap=True)
        col += 2

    for label, bg in extra_cols:
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        cell = ws.cell(row=3, column=col, value=label)
        cell.font      = _font(bold=True, size=12, color=C["white"])
        cell.fill      = _fill(bg)
        cell.alignment = _align(wrap=True)
        col += 2

    _set_row_height(ws, 3, 50)

    # Top 5 Gainers / Losers
    ws.merge_cells("A5:H5")
    _write(ws, 5, 1, "TOP 5 GAINERS", bold=True, fg=C["white"], bg=C["green_str"], size=12)
    ws.merge_cells("I5:P5")
    _write(ws, 5, 9, "TOP 5 LOSERS", bold=True, fg=C["white"], bg=C["red_str"], size=12)

    gl_cols = ["Symbol", "Sector", "Close", "Chg ₹", "Chg %", "Del %", "OI Chg%", "Signal"]
    _header_row(ws, 6, gl_cols, bg=C["header_bg"])
    _header_row(ws, 6, gl_cols, bg=C["header_bg"])
    for i, c in enumerate(gl_cols, 9):
        _write(ws, 6, i, c, bold=True, fg=C["white"], bg=C["header_bg"], border=True)

    gainers = signals.nlargest(5, "chg_pct")
    losers  = signals.nsmallest(5, "chg_pct")

    for r, (_, row) in enumerate(gainers.iterrows(), 7):
        row_bg = C["light_grey"] if r % 2 == 0 else C["white"]
        for ci, val in enumerate([
            row["symbol"], row["sector"], f"{row['close']:.2f}",
            f"{row['chg']:+.2f}", f"{row['chg_pct']:+.2f}%",
            f"{row['deliv_pct']:.2f}", f"{row['oi_chg_pct']:+.2f}%",
        ], 1):
            _write(ws, r, ci, val, bg=row_bg, border=True)
        _signal_cell(ws, r, 8, row["signal_label"])

    for r, (_, row) in enumerate(losers.iterrows(), 7):
        row_bg = C["light_grey"] if r % 2 == 0 else C["white"]
        for ci, val in enumerate([
            row["symbol"], row["sector"], f"{row['close']:.2f}",
            f"{row['chg']:+.2f}", f"{row['chg_pct']:+.2f}%",
            f"{row['deliv_pct']:.2f}", f"{row['oi_chg_pct']:+.2f}%",
        ], 9):
            _write(ws, r, ci, val, bg=row_bg, border=True)
        _signal_cell(ws, r, 16, row["signal_label"])

    # Sector-wise analysis
    row_start = 13
    ws.merge_cells(f"A{row_start}:J{row_start}")
    _write(ws, row_start, 1, "SECTOR-WISE ANALYSIS", bold=True, size=13,
           fg=C["white"], bg=C["bg_mid"])

    sec_headers = ["Sector", "Stocks", "Avg Del%", "Del vs Avg4",
                   "Avg Price Chg%", "Bullish", "Bearish", "Avg OI Chg%",
                   "Net FII", "Sector Signal"]
    _header_row(ws, row_start + 1, sec_headers)

    for r, (_, row) in enumerate(sector.iterrows(), row_start + 2):
        row_bg = C["row_alt"] if r % 2 == 0 else C["white"]
        for ci, val in enumerate([
            row.get("sector", ""),
            int(row.get("stocks", 0)),
            f"{row.get('avg_del_pct', 0):.2f}",
            f"{row.get('avg_del_pct', 0) - 40:.2f}",  # vs 40% benchmark
            f"{row.get('avg_chg_pct', 0):.2f}%",
            int(row.get("bullish", 0)),
            int(row.get("bearish", 0)),
            f"{row.get('avg_oi_chg', 0):+.2f}%",
            "—",
        ], 1):
            _write(ws, r, ci, val, bg=row_bg, border=True)
        _signal_cell(ws, r, 10, row.get("sector_signal", "⬜ Neutral"))

    _set_col_widths(ws, [12, 13, 9, 9, 9, 9, 9, 9, 11, 11, 12, 9, 9, 9, 9, 14])


# ─── Sheet 2: Stock Detail ────────────────────────────────────────────────────

def _sheet_stock_detail(wb: Workbook, signals: pd.DataFrame,
                        hist: pd.DataFrame, market_date: date, run_time: str):
    ws = wb.create_sheet("📋 Stock Detail")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:Z1")
    title = f"  NSE NIFTY 50  ·  STOCK DETAIL  ·  {market_date.strftime('%d %b %Y')}  {run_time}"
    _write(ws, 1, 1, title, bold=True, size=14, fg=C["gold"], bg=C["bg_dark"], h="left")

    # Group headers row 3
    groups = [
        ("A3:B3", "IDENTITY"),
        ("C3:E3", "TODAY PRICE"),
        ("F3:I3", "PREV 4 CLOSES"),
        ("J3:N3", "DELIVERY %"),
        ("O3:P3", "DEL DIFF"),
        ("Q3:R3", "VOLUME"),
        ("S3:T3", "OI FUTURES"),
        ("U3:V3", "52W"),
        ("W3:X3", "SIGNAL"),
        ("Y3:Z3", "DATES"),
    ]
    for cell_range, label in groups:
        ws.merge_cells(cell_range)
        start_col = ws[cell_range.split(":")[0]].column
        _write(ws, 3, start_col, label, bold=True, fg=C["white"],
               bg=C["bg_mid"], border=True)

    # Sub-headers row 4
    sub = [
        "SYMBOL", "SECTOR",
        "Open", "Close\nToday", "Chg\n₹ / %",
        "Close\nD-1", "Close\nD-2", "Close\nD-3", "Close\nD-4",
        "Del%\nToday", "Del%\nD-1", "Del%\nD-2", "Del%\nD-3", "Del%\nD-4",
        "Diff\nvs D-1", "Diff\nvs Avg4",
        "Vol\nToday", "Vol\nRatio",
        "OI\nChg", "OI\nChg%",
        "52W\nHigh", "52W\nLow",
        "Signal", "Exp\nMove",
        "Date\nToday", "Date\nD-1",
    ]
    _header_row(ws, 4, sub)
    _set_row_height(ws, 4, 36)

    # Build historical lookup
    hist_map = {}
    if not hist.empty:
        hist_sorted = hist.sort_values("date", ascending=False)
        for sym, grp in hist_sorted.groupby("symbol"):
            hist_map[sym] = grp.reset_index(drop=True)

    for r, (_, row) in enumerate(signals.iterrows(), 5):
        sym  = row["symbol"]
        h    = hist_map.get(sym, pd.DataFrame())
        bg   = C["row_alt"] if r % 2 == 0 else C["white"]

        def _hv(col, idx, default=0):
            try:
                return float(h.iloc[idx][col]) if len(h) > idx else default
            except Exception:
                return default

        prev_closes  = [_hv("close",     i, row["prev_close"]) for i in range(4)]
        prev_dels    = [_hv("deliv_pct", i, row["deliv_pct"])  for i in range(4)]
        avg4_del     = sum(prev_dels) / 4 if prev_dels else row["deliv_pct"]
        avg5_vol     = h["volume"].head(5).mean() if len(h) >= 2 else row["volume"]
        vol_ratio    = round(row["volume"] / avg5_vol, 2) if avg5_vol else 1.0
        high52       = h["high"].max()  if len(h) >= 2 else row.get("high", 0)
        low52        = h["low"].min()   if len(h) >= 2 else row.get("low",  0)
        prev_date    = str(h.iloc[0]["date"]) if len(h) >= 1 else ""

        chg_str = f"{row['chg']:+.2f}\n{row['chg_pct']:+.2f}%"

        vals = [
            sym, row["sector"],
            round(row["open"], 2), round(row["close"], 2), chg_str,
            *[round(c, 2) for c in prev_closes],
            round(row["deliv_pct"], 2), *[round(d, 2) for d in prev_dels],
            round(row["deliv_pct"] - prev_dels[0], 2) if prev_dels else 0,
            round(row["deliv_pct"] - avg4_del, 2),
            int(row["volume"]), vol_ratio,
            int(row["oi_chg"]), round(row["oi_chg_pct"], 2),
            round(high52, 2), round(low52, 2),
            row["signal_label"], "",
            str(market_date), prev_date,
        ]

        for ci, val in enumerate(vals, 1):
            if ci == 23:  # Signal column
                _signal_cell(ws, r, ci, str(val))
            else:
                _write(ws, r, ci, val, bg=bg, border=True, wrap=(ci == 5))

    _set_col_widths(ws, [
        12, 13, 9, 10, 12, 9, 9, 9, 9,
        9, 9, 9, 9, 9, 9, 9,
        12, 8, 10, 8, 9, 9,
        18, 12, 12, 12,
    ])


# ─── Sheet 3: FII-DII ─────────────────────────────────────────────────────────

def _sheet_fii(wb: Workbook, fii: pd.Series, fii_hist: pd.DataFrame,
               market_date: date, run_time: str):
    ws = wb.create_sheet("💹 FII-DII")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B1:E1")
    _write(ws, 1, 2, f"FII / DII CASH MARKET  ·  {market_date.strftime('%d %b %Y')}  {run_time}",
           bold=True, size=14, fg=C["gold"], bg=C["bg_dark"])

    _header_row(ws, 3, ["", "Category", "Buy (₹ Cr)", "Sell (₹ Cr)", "Net (₹ Cr)"], bg=C["header_bg"])

    fii_net = float(fii.get("fii_net", 0)) if not fii.empty else 0
    dii_net = float(fii.get("dii_net", 0)) if not fii.empty else 0

    rows_data = [
        ("FII / FPI",
         fii.get("fii_buy", 0), fii.get("fii_sell", 0), fii_net),
        ("DII",
         fii.get("dii_buy", 0), fii.get("dii_sell", 0), dii_net),
    ]
    for r_idx, (cat, buy, sell, net) in enumerate(rows_data, 4):
        bg = C["row_alt"] if r_idx % 2 == 0 else C["white"]
        _write(ws, r_idx, 2, cat,  bold=True, bg=bg, border=True, h="left")
        _write(ws, r_idx, 3, round(float(buy),  2), bg=bg, border=True)
        _write(ws, r_idx, 4, round(float(sell), 2), bg=bg, border=True)
        net_bg = "E8F5E9" if float(net) >= 0 else "FFEBEE"
        _write(ws, r_idx, 5, round(float(net),  2), bold=True, bg=net_bg, border=True)

    # Historical trend
    if not fii_hist.empty:
        ws.merge_cells("B7:E7")
        _write(ws, 7, 2, "RECENT FII/DII TREND", bold=True, size=12,
               fg=C["white"], bg=C["bg_mid"])
        _header_row(ws, 8, ["", "Date", "FII Net", "DII Net", "Combined Net"])
        for r_idx, (_, hr) in enumerate(fii_hist.head(10).iterrows(), 9):
            bg = C["row_alt"] if r_idx % 2 == 0 else C["white"]
            combined = float(hr.get("fii_net", 0)) + float(hr.get("dii_net", 0))
            _write(ws, r_idx, 2, str(hr.get("date", "")), bg=bg, border=True)
            _write(ws, r_idx, 3, round(float(hr.get("fii_net", 0)), 2), bg=bg, border=True)
            _write(ws, r_idx, 4, round(float(hr.get("dii_net", 0)), 2), bg=bg, border=True)
            net_bg = "E8F5E9" if combined >= 0 else "FFEBEE"
            _write(ws, r_idx, 5, round(combined, 2), bold=True, bg=net_bg, border=True)

    _set_col_widths(ws, [2, 18, 14, 14, 14])


# ─── Sheet 4: Signals ─────────────────────────────────────────────────────────

def _sheet_signals(wb: Workbook, signals: pd.DataFrame, market_date: date):
    ws = wb.create_sheet("📈 Signals")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:M1")
    _write(ws, 1, 1, f"  SIGNAL DETAIL  ·  {market_date.strftime('%d %b %Y')}",
           bold=True, size=14, fg=C["gold"], bg=C["bg_dark"], h="left")

    headers = ["Symbol", "Sector", "Score", "Signal", "Rec",
               "S:Del", "S:Price", "S:OI", "S:Vol", "S:FII",
               "OI Signal", "Del%", "OI Chg%"]
    _header_row(ws, 3, headers)

    for r, (_, row) in enumerate(signals.iterrows(), 4):
        bg = C["row_alt"] if r % 2 == 0 else C["white"]
        vals = [
            row["symbol"], row["sector"], row["score"],
        ]
        for ci, val in enumerate(vals, 1):
            _write(ws, r, ci, val, bg=bg, border=True)
        _signal_cell(ws, r, 4, row["signal_label"])
        for ci, val in enumerate([
            row["recommendation"],
            row.get("s_delivery",  0), row.get("s_price",   0),
            row.get("s_oi",        0), row.get("s_volume",  0),
            row.get("s_fii",       0), row.get("oi_signal", ""),
            round(row.get("deliv_pct",    0), 2),
            round(row.get("oi_chg_pct",   0), 2),
        ], 5):
            _write(ws, r, ci, val, bg=bg, border=True)

    _set_col_widths(ws, [12, 13, 8, 18, 20, 7, 8, 7, 7, 7, 16, 8, 9])


# ─── Sheet 5: Logs ────────────────────────────────────────────────────────────

def _sheet_logs(wb: Workbook, run_log: pd.DataFrame):
    ws = wb.create_sheet("🗒️ Logs")
    ws.sheet_view.showGridLines = False
    _header_row(ws, 1, ["Run Time", "Market Date", "Status", "Message"])
    for r, (_, row) in enumerate(run_log.iterrows(), 2):
        bg = "E8F5E9" if str(row.get("status","")).upper() == "SUCCESS" else "FFEBEE"
        for ci, col in enumerate(["run_time","market_date","status","message"], 1):
            _write(ws, r, ci, str(row.get(col, "")), bg=bg, border=True, h="left")
    _set_col_widths(ws, [22, 14, 12, 60])


# ─── Main entry point ─────────────────────────────────────────────────────────

def generate_excel(
    signals_df: pd.DataFrame,
    fii_series: pd.Series,
    sector_df:  pd.DataFrame,
    hist_df:    pd.DataFrame,
    fii_hist:   pd.DataFrame,
    run_log_df: pd.DataFrame,
    market_date: date,
    output_path: str,
):
    """Generate the full Excel dashboard."""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)   # remove default Sheet

    run_time = datetime.now().strftime("%I:%M %p")

    _sheet_dashboard(wb, signals_df, fii_series, sector_df, market_date, run_time)
    _sheet_stock_detail(wb, signals_df, hist_df, market_date, run_time)
    _sheet_fii(wb, fii_series, fii_hist, market_date, run_time)
    _sheet_signals(wb, signals_df, market_date)
    _sheet_logs(wb, run_log_df)

    wb.save(output_path)
    log.info(f"Excel saved: {output_path}")
